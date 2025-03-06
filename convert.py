import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from operator import itemgetter
import json
import sys
import logging

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 설정 상수
DATETIME_FORMAT = '%H:%M:%S'
LATE_THRESHOLD = '08:30:59'
WEEKDAY_MAP = {'Monday': '월', 'Tuesday': '화', 'Wednesday': '수', 'Thursday': '목', 'Friday': '금'}

# 스타일 설정
STYLES = {
    'late_fill': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
    'late_font': Font(color='FF0000'),
    'estimated_fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
    'default_font': Font(name='맑은 고딕', size=12),
    'alignment': Alignment(horizontal='center', vertical='center'),
    'monday_border': Border(top=Side(border_style='double'))
}

def add_12_hours(time_str):
    """오후 시간을 12시간 추가하여 변환"""
    time_obj = datetime.strptime(time_str, DATETIME_FORMAT)
    new_time_obj = time_obj + timedelta(hours=12)
    return new_time_obj.strftime(DATETIME_FORMAT)

def get_all_dates(data_dict):
    """최소/최대 날짜 사이 모든 날짜 반환"""
    sorted_dates = sorted(data_dict.keys(), key=lambda x: datetime.strptime(x, '%Y-%m-%d'))
    min_date = datetime.strptime(sorted_dates[0], '%Y-%m-%d')
    max_date = datetime.strptime(sorted_dates[-1], '%Y-%m-%d')
    return [(min_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((max_date - min_date).days + 1)]

def parse_datetime(datetime_str):
    """날짜/시간 문자열 파싱"""
    try:
        parts = datetime_str.split(' ')
        if len(parts) == 3:
            date, ampm, time = parts
            if ampm == '오후':
                time = add_12_hours(time)
        elif len(parts) == 2:
            date, time = parts
        else:
            raise ValueError(f"Invalid format: {datetime_str}")
        return date, time
    except ValueError as e:
        logger.error(f"Error parsing datetime: {e}, skipping: {datetime_str}")
        return None, None

def load_data(file_path):
    """엑셀 파일에서 데이터 로드"""
    try:
        excel_df = pd.read_excel(file_path, skiprows=2)
        rename_df = excel_df.rename(columns={
            "인증일시": "date_Attestation", "요일": "str_Week", "인증번호": "str_tmId",
            "사원번호": "str_workempNum", "이름": "str_workempName", "리더기 장소": "str_accTerminalPlace",
            "인증모드": "str_Mode", "인증상태": "str_ValidationStatus", "인증방법": "str_Certificate",
            "부서": "str_workempPostName", "직위": "str_workempPositionName",
            "타임테이블": "str_workUserTimetableName", "직원상태": "str_emptmAdmin"
        })
        json_data = rename_df.to_json(orient='records', force_ascii=False)
        logger.info("Data loaded successfully from %s", file_path)
        return json.loads(json_data)
    except Exception as e:
        logger.error("Failed to load data from %s: %s", file_path, e)
        return None

def process_attendance(work_data):
    """출퇴근 데이터 처리"""
    attendance_dict = {}
    for entry in work_data:
        datetime_str, mode, name = itemgetter('date_Attestation', 'str_Mode', 'str_workempName')(entry)
        date, time = parse_datetime(datetime_str)
        if date is None or time is None:
            continue

        if date not in attendance_dict:
            attendance_dict[date] = {}
        if name not in attendance_dict[date]:
            attendance_dict[date][name] = {'출근': '', '퇴근': '', 'times': [], '출근_추정': False, '퇴근_추정': False}

        attendance_dict[date][name]['times'].append(time)
        if mode in ['출근', '퇴근']:
            current_time = attendance_dict[date][name][mode]
            if mode == '출근' and (not current_time or datetime.strptime(time, DATETIME_FORMAT) < datetime.strptime(current_time, DATETIME_FORMAT)):
                attendance_dict[date][name][mode] = time
            elif mode == '퇴근' and (not current_time or datetime.strptime(time, DATETIME_FORMAT) > datetime.strptime(current_time, DATETIME_FORMAT)):
                attendance_dict[date][name][mode] = time

    # 추정값 처리
    for date in attendance_dict:
        for name in attendance_dict[date]:
            if attendance_dict[date][name]['times']:
                times = sorted(attendance_dict[date][name]['times'])
                if not attendance_dict[date][name]['출근']:
                    attendance_dict[date][name]['출근'] = times[0]
                    attendance_dict[date][name]['출근_추정'] = True
                if not attendance_dict[date][name]['퇴근']:
                    attendance_dict[date][name]['퇴근'] = times[-1]
                    attendance_dict[date][name]['퇴근_추정'] = True
    return attendance_dict

def generate_rows(attendance_dict, employee_names):
    """엑셀에 기록할 행 데이터 생성"""
    rows = []
    all_dates = get_all_dates(attendance_dict)
    for date in all_dates:
        weekday = datetime.strptime(date, '%Y-%m-%d').strftime('%A')
        if weekday in ['Saturday', 'Sunday']:
            continue

        weekday_kr = WEEKDAY_MAP[weekday]
        row = [datetime.strptime(date, '%Y-%m-%d').strftime('%Y/%m/%d'), weekday_kr, '']
        if date in attendance_dict:
            for employee in employee_names:
                if employee in attendance_dict[date]:
                    row.append((attendance_dict[date][employee]['출근'], attendance_dict[date][employee]['출근_추정']))
                    row.append((attendance_dict[date][employee]['퇴근'], attendance_dict[date][employee]['퇴근_추정']))
                else:
                    row.append(('', False))
                    row.append(('', False))
        else:
            for _ in employee_names:
                row.append(('', False))
                row.append(('', False))
        rows.append(row)
    return rows

def setup_workbook(employee_names):
    """엑셀 워크북 초기 설정"""
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'], sheet['B1'], sheet['C1'] = '날짜', '요일', '비고'
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    sheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    for idx, col in enumerate(employee_names):
        start_column = 4 + idx * 2
        sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column + 1)
        sheet[get_column_letter(start_column) + '1'] = col
        sheet[get_column_letter(start_column) + '2'] = '출근'
        sheet[get_column_letter(start_column + 1) + '2'] = '퇴근'

    sheet.column_dimensions['A'].width = 16
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 30
    for col_idx in range(4, 4 + len(employee_names) * 2):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 10.5
    sheet.freeze_panes = 'C3'
    return workbook

def apply_styles(sheet, rows):
    """엑셀 스타일 적용"""
    for row in sheet.rows:
        for cell in row:
            cell.font = STYLES['default_font']
            cell.alignment = STYLES['alignment']

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, max_col=sheet.max_column):
        if row[1].value == '월':
            for cell in row:
                cell.border = STYLES['monday_border']
        for idx, cell in enumerate(row[3:], start=3):  # 직원 데이터만
            if cell.value:
                row_idx, col_idx = cell.row - 3, cell.column - 1
                is_estimated = isinstance(rows[row_idx][col_idx], tuple) and rows[row_idx][col_idx][1]
                if idx % 2 == 1:  # 출근 열
                    try:
                        time_val = datetime.strptime(cell.value, DATETIME_FORMAT)
                        if time_val > datetime.strptime(LATE_THRESHOLD, DATETIME_FORMAT):
                            cell.fill = STYLES['late_fill']
                            cell.font = STYLES['late_font']
                    except ValueError:
                        pass
                if is_estimated:
                    cell.fill = STYLES['estimated_fill']

def convert(file_path, employee_order=None):
    """메인 변환 함수"""
    work_data = load_data(file_path)
    if not work_data:
        return None

    attendance_dict = process_attendance(work_data)
    employee_names = employee_order if employee_order else sorted(set(entry['str_workempName'] for entry in work_data))
    rows = generate_rows(attendance_dict, employee_names)
    workbook = setup_workbook(employee_names)

    sheet = workbook.active
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx).value = value[0] if isinstance(value, tuple) else value

    apply_styles(sheet, rows)
    logger.info("Workbook conversion completed")
    return workbook

# 사용 예시
if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "input.xlsx"  # Excel 파일 경로
    # 하드코딩된 employee_order 사용
    employee_order = [
        "강희경(Sophie)", "김민경(Ari)", "김민규(Arthur)", "김성준(Alex)", "김영석(Ethan)",
        "김정한(Hans)", "박주헌(Stark)", "성영아(Amy)", "양은영(Ella)", "오준석(Alex)",
        "유주영(Roxie)", "정기철(Roy)", "정대웅(Henry)", "최정원(Jen)", "박병건(Ben)",
        "서이현(Zoe)", "정재윤(Rio)"
    ]
    workbook = convert(file_path, employee_order)
    if workbook:
        logger.info("Conversion completed successfully")
    